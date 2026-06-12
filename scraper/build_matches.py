#!/usr/bin/env python3
"""
Genera data/matches.json a partir del Excel de pronósticos.
Uso: python scraper/build_matches.py [ruta_al_xlsx]
Por defecto busca ../Pollas_Mundial_2026.xlsx
"""
import openpyxl, json, os, sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = sys.argv[1] if len(sys.argv) > 1 else os.path.join(os.path.dirname(ROOT), "Pollas_Mundial_2026.xlsx")
OUT = os.path.join(ROOT, "data", "matches.json")

MESES = {'ene':1,'feb':2,'mar':3,'abr':4,'may':5,'jun':6,'jul':7,'ago':8,'sep':9,'oct':10,'nov':11,'dic':12}
# columnas 1-based del Excel -> id de polla
COL2POLLA = {7:'pollapato', 8:'polla26', 9:'triplea', 10:'triplea2', 11:'kicktipp'}

def parse_score(s):
    if s is None: return None
    s = str(s).replace('–','-').strip(); p = s.split('-')
    if len(p) != 2: return None
    try: return [int(p[0]), int(p[1])]
    except: return None

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Pollas"]
    matches = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        num = row[0]
        if num is None: continue
        fecha, hora, grupo, local, visit = row[1], row[2], row[3], row[4], row[5]
        sede = row[11]
        if not (fecha and local): continue
        d, mes = str(fecha).split(); month = MESES[mes.lower()[:3]]; day = int(d)
        hh, mm = str(hora or '00:00').split(':')
        iso = f"2026-{month:02d}-{day:02d}T{int(hh):02d}:{int(mm):02d}:00-05:00"
        preds = {pid: parse_score(row[c-1]) for c, pid in COL2POLLA.items()}
        matches.append({"n": int(num), "kickoff": iso, "date": f"2026-{month:02d}-{day:02d}",
            "time": f"{int(hh):02d}:{int(mm):02d}", "group": grupo, "home": local, "away": visit,
            "venue": sede, "predictions": preds, "result": None})
    out = {"generated": "2026", "timezone": "America/Panama (UTC-5)", "count": len(matches), "matches": matches}
    json.dump(out, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    print(f"OK: {len(matches)} partidos -> {OUT}")

if __name__ == "__main__":
    main()
